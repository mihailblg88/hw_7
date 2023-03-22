import PyPDF2
import csv
import os
import requests
import zipfile
from zipfile import ZipFile
import openpyxl
from openpyxl.reader.excel import load_workbook

name_zip = 'all_files.zip'

url_csv = 'https://file-examples.com/wp-content/uploads/2017/02/file_example_CSV_5000.csv'
url_xlsx = 'https://file-examples.com/wp-content/uploads/2017/02/file_example_XLSX_50.xlsx'
url_pdf = 'https://www.orimi.com/pdf-test.pdf'

filename_csv = os.path.basename(url_csv)
filename_xlsx = os.path.basename(url_xlsx)
filename_pdf = os.path.basename(url_pdf)


def download_files(file_csv, file_xlsx, file_pdf):
    try:
        os.mkdir('../all_files')
    except:
        print('Создана папка all_files')

    response = requests.get(url_csv)
    with open(f'../all_files/{file_csv}', 'wb') as csv:
        csv.write(response.content)

    response = requests.get(url_xlsx)
    with open(f'../all_files/{file_xlsx}', 'wb') as xlsx:
        xlsx.write(response.content)

    response = requests.get(url_pdf)
    with open(f'../all_files/{file_pdf}', 'wb') as pdf:
        pdf.write(response.content)


def zip_make(name_zip, file_pdf, file_xlsx, file_csv):
    try:
        os.mkdir('../zip_files')
    except:
        print('Создана папка zip_files')

    with zipfile.ZipFile(f'../zip_files/{name_zip}', 'w') as myzip:
        myzip.write(f'all_files/{file_csv}', filename_csv)
        myzip.write(f'all_files/{file_xlsx}', filename_xlsx)
        myzip.write(f'all_files/{file_pdf}', filename_pdf)


def csv_size(name_zip, file_csv):
    with zipfile.ZipFile(f'../zip_files/{name_zip}', 'r') as myzip:
        zip_size_csv = myzip.getinfo(filename_csv).file_size
    file_size_csv = os.path.getsize(f'../all_files/{file_csv}')
    return zip_size_csv == file_size_csv


def xlsx_size(name_zip, file_xlsx):
    with zipfile.ZipFile(f'../zip_files/{name_zip}', 'r') as myzip:
        zip_size_xls = myzip.getinfo(filename_xlsx).file_size
    file_size_xlsx = os.path.getsize(f'../all_files/{file_xlsx}')
    return zip_size_xls == file_size_xlsx


def pdf_size(name_zip, file_pdf):
    with zipfile.ZipFile(f'../zip_files/{name_zip}', 'r') as myzip:
        zip_size_pdf = myzip.getinfo(filename_pdf).file_size
    file_size_pdf = os.path.getsize(f'../all_files/{file_pdf}')
    return zip_size_pdf == file_size_pdf


def check_csv_row(name_zip, file_csv):
    row_count_arc = 0

    with open(f'../all_files/{file_csv}', "r") as csv_file:
        file = csv.reader(csv_file, delimiter=";")
        row_count = sum(1 for row in file)

    with ZipFile(f'../zip_files/{name_zip}') as myzip:
        with myzip.open(f"{file_csv}", "r") as myfilezip:
            for i in myfilezip:
                row_count_arc += 1

    return row_count == row_count_arc


def check_xlsx_row(name_zip, file_xlsx):
    file = load_workbook(f'../all_files/{file_xlsx}')
    with open(f'../all_files/{file_xlsx}', 'r') as pdf_file:
        sheet = file.active
        xlsx_row_file = sheet.max_row

    with zipfile.ZipFile(f'../zip_files/{name_zip}', 'r') as myzip:
        xlsxfile = myzip.open(file_xlsx)
        xlsx_reader = openpyxl.load_workbook(xlsxfile)
        sheet = xlsx_reader.active
        xlsx_rows_archive = sheet.max_row

    return xlsx_row_file == xlsx_rows_archive


def check_pdf_page(name_zip, file_pdf):
    with open(f'../all_files/{file_pdf}', "rb") as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        pdf_page_file = len(pdf_reader.pages)

    with zipfile.ZipFile(f'../zip_files/{name_zip}', 'r') as myzip:
        pdffile = myzip.open(file_pdf)
        pdf_reader = PyPDF2.PdfReader(pdffile)
        pdf_page_archive = len(pdf_reader.pages)

    return pdf_page_file == pdf_page_archive


def test_hm_7():
    download_files(filename_csv, filename_xlsx, filename_pdf)
    zip_make(name_zip, filename_csv, filename_xlsx, filename_pdf)

    assert pdf_size(name_zip, filename_csv)
    assert xlsx_size(name_zip, filename_xlsx)
    assert csv_size(name_zip, filename_pdf)

    assert check_pdf_page(name_zip, filename_csv)
    assert check_xlsx_row(name_zip, filename_xlsx)
    assert check_csv_row(name_zip, filename_pdf)
