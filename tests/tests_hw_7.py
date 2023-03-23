import PyPDF2
import csv
import os
import requests
import zipfile
from zipfile import ZipFile
import openpyxl
from openpyxl.reader.excel import load_workbook

name_zip = 'zip_files.zip'


url_pdf = 'https://freetestdata.com/wp-content/uploads/2021/09/Free_Test_Data_100KB_PDF.pdf'
url_xlsx = "https://freetestdata.com/wp-content/uploads/2021/09/Free_Test_Data_100KB_XLSX.xlsx"
url_csv = 'https://freetestdata.com/wp-content/uploads/2021/09/Free_Test_Data_200KB_CSV-1.csv'

filename_csv = os.path.basename(url_csv)
filename_xlsx = os.path.basename(url_xlsx)
filename_pdf = os.path.basename(url_pdf)


def download_files(file_pdf, file_xlsx, file_csv):
    try:
        os.mkdir('all_files')
    except:
        print('Создана папка all_files')

    response = requests.get(url_pdf)
    with open(f'all_files/{file_pdf}', 'wb') as pdf:
        pdf.write(response.content)

    response = requests.get(url_xlsx)
    with open(f'all_files/{file_xlsx}', 'wb') as xlsx:
        xlsx.write(response.content)

    response = requests.get(url_csv)
    with open(f'all_files/{file_csv}', 'wb') as csv:
        csv.write(response.content)


def zip_make(name_zip, file_pdf, file_xlsx, file_csv):
    try:
        os.mkdir('resources')
    except:
        print('Создана папка resources')

    with zipfile.ZipFile(f'resources/{name_zip}', 'w') as myzip:
        myzip.write(f'all_files/{file_pdf}', filename_pdf)
        myzip.write(f'all_files/{file_xlsx}', filename_xlsx)
        myzip.write(f'all_files/{file_csv}', filename_csv)


def check_pdf_size(name_zip, file_pdf):
    with zipfile.ZipFile(f'resources/{name_zip}', 'r') as myzip:
        zip_size_pdf = myzip.getinfo(filename_pdf).file_size
    file_size_pdf = os.path.getsize(f'all_files/{file_pdf}')
    return zip_size_pdf == file_size_pdf


def xlsx_size(name_zip, file_xlsx):
    with zipfile.ZipFile(f'resources/{name_zip}', 'r') as myzip:
        zip_size_xls = myzip.getinfo(filename_xlsx).file_size
    file_size_xlsx = os.path.getsize(f'all_files/{file_xlsx}')
    return zip_size_xls == file_size_xlsx


def csv_size(name_zip, file_csv):
    with zipfile.ZipFile(f'resources/{name_zip}', 'r') as myzip:
        zip_size_csv = myzip.getinfo(filename_csv).file_size
    file_size_csv = os.path.getsize(f'all_files/{file_csv}')
    return zip_size_csv == file_size_csv


def pdf_size(name_zip, file_pdf):
    with open(f'all_files/{file_pdf}', "rb") as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        pdf_page_file = len(pdf_reader.pages)

    with zipfile.ZipFile(f'resources/{name_zip}', 'r') as myzip:
        pdffile = myzip.open(file_pdf)
        pdf_reader = PyPDF2.PdfReader(pdffile)
        pdf_page_archive = len(pdf_reader.pages)

    return pdf_page_file == pdf_page_archive


def check_xlsx(name_zip, file_xlsx):
    file = load_workbook(f'all_files/{file_xlsx}')
    with open(f'all_files/{file_xlsx}', 'r') as xlsx_file:
        sheet = file.active
        xlsx_row_file = sheet.max_row

    with zipfile.ZipFile(f'resources/{name_zip}', 'r') as myzip:
        xlsxfile = myzip.open(file_xlsx)
        xlsx_reader = openpyxl.load_workbook(xlsxfile)
        sheet = xlsx_reader.active
        xlsx_rows_archive = sheet.max_row

    return xlsx_row_file == xlsx_rows_archive


def check_csv(name_zip, file_csv):
    row_count_arc = 0

    with open(f'all_files/{file_csv}', "r") as csv_file:
        file = csv.reader(csv_file, delimiter=";")
        row_count = sum(1 for row in file)

    with ZipFile(f'resources/{name_zip}') as myzip:
        with myzip.open(f"{file_csv}", "r") as myfilezip:
            for i in myfilezip:
                row_count_arc += 1

    return row_count == row_count_arc


def test_hm7():
    download_files(filename_pdf, filename_xlsx, filename_csv)
    zip_make(name_zip, filename_pdf, filename_xlsx, filename_csv)

    assert pdf_size(name_zip, filename_pdf)
    assert xlsx_size(name_zip, filename_xlsx)
    assert csv_size(name_zip, filename_csv)

    assert pdf_size(name_zip, filename_pdf)
    assert check_xlsx(name_zip, filename_xlsx)
    assert check_csv(name_zip, filename_csv)